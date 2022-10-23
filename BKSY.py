from openpyxl import Workbook
import requests
import re
import datetime
import openpyxl
url = 'https://www.bkjx.sdu.edu.cn/sanji_list.jsp?urltype=tree.TreeTempUrl&wbtreeid=1010'
headers = {
    'cookie': '_zap=6d20b1b3-0e83-4a23-86f8-a1a613065e2b; d_c0="ADCfiUSeQBWPTmcpyWjGq2_W-FiigLWZ2hU=|1657893709"; gdxidpyhxdE=n31M0D3qe4ne5k6SSYZO9O0WzZ0ckhGgIAqjB4ZunhaoHJ5eiW1LHMrPw2o3G%5C3B94Amy0PhOZu6Orpgp9zNUglLauXwWu6ob6z98XcPO%2Fqtts3%2FQuUfUT6JOTkAl8w1uUOAnjjMrN5hvEnyvsPjHhq%2FKfvffS%2FO5qnrWuXi8JqhJRbW%3A1657894613843; _9755xjdesxxd_=32; YD00517437729195%3AWM_NI=Pq7KzWorpm7mhgp9ml1HHrNOgnwSTsHIa5CJ95oVSVOcOgz3%2F%2FiqHe6prkvDpLV6InuvhCkO3d46UEltKIb%2ByfK6dkZ%2B0rwGoGbZNYFMSsrjlQ%2BAGYXg4psVsUedbs1aM2w%3D; YD00517437729195%3AWM_NIKE=9ca17ae2e6ffcda170e2e6eeb1fc4f90bda1d3d97e8bbc8fa2d44f829e9bb1c859928a83b5d06dedeeb9b8d12af0fea7c3b92a94aff9acf95dadaec08eb667acb9c0daf068f3988893cd45a5b2b9b4d7798c97afd6c242fbeba1b7d25e8dbd8188f939edf1a792cb6ff6b0be8cf550b4949ad7cc66b7aaa8a4ea45f1bcfcd7d84fb0ac85aae121bbba85d7e7619af50093f64dad99aebbcd5386939d94f162b08fe5a2c6539cb9fc91f25eb18bab85e179f3bf83d3f637e2a3; YD00517437729195%3AWM_TID=%2BW0lvvbmt7RAREREFEeBXRFMWM062wCR; captcha_session_v2=2|1:0|10:1657893716|18:captcha_session_v2|88:WFlMeDR3TTNnSGIxelhiOWM2VWhtVHNNaHVsdFIvMFJFMTJDemhMUDl3ZEdVeWNJNGVLV0lWZWg5RDJMY29vag==|d5e1889c5712bd714d1c31a39a489cf6f58be0642702be6f6b46f79758866787; z_c0=2|1:0|10:1657893721|4:z_c0|92:Mi4xYThhZkNnQUFBQUFBTUotSlJKNUFGUmNBQUFCZ0FsVk5XY0ctWXdEcUkxQmctT2tCMWlxTzl4TUZUcFdDN3E1eldR|aa82f3dd2a774fa57cc87dcfb07be71040e3a53dd1919313e0a13a50cd544b97; _xsrf=83abd56b-1867-4269-bb57-960cfd5c91a3; q_c1=357fa7b4d4184eaabd568cee051209e2|1658716117000|1658716117000; Hm_lvt_98beee57fd2ef70ccdd5ca52b9740c49=1658040783,1658128773,1658388682,1658716119; Hm_lpvt_98beee57fd2ef70ccdd5ca52b9740c49=1658716119; ariaDefaultTheme=undefined; NOT_UNREGISTER_WAITING=1; BAIDU_SSP_lcr=https://www.baidu.com/link?url=k6d8dLG71FQpb0v6O3n1IhpIx5uPWk8oDctz-vv11ki&wd=&eqid=9bf2497900047a050000000462ddffd3; SESSIONID=BmmDYzkeV0798gpIq5AEdXMDOfjNQs5GnI6luIlzjHB; JOID=Ul8XC0Jlj6ZJkJcMOW2RO3bSAUIuIOfBDMDqYlY6uJR6qfl7db5_cCOelQMyNcV7d5N13UO9qZ1Ax4lO1mvNbnk=; osd=Vl0XAU5hjaZDnJMOOWedP3TSC04qIufLAMToYlw2vJZ6o_V_d751fCeclQk-Mcd7fZ9x30O3pZlCx4NC0mnNZHU=; tst=h; KLBRSID=37f2e85292ebb2c2ef70f1d8e39c2b34|1658716150|1658716117',
    'referer': 'https://www.bkjx.sdu.edu.cn/sanji_list.jsp?urltype=tree.TreeTempUrl&wbtreeid=1010',
    'user-agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/103.0.0.0 Safari/537.36'
}
page_data = requests.get(url=url, headers=headers).text
biaoti = re.findall('<div style="float:left">.*?title="(.*?)" style', page_data, re.S)
lianjie=re.findall('<div style="float:left">.*?href="(.*?)" target',page_data,re.S)
shijian=re.findall('<div style="float:right;">(.*?)</div>',page_data,re.S)
print(lianjie)
ti = datetime.datetime.now()
time_hot = str(ti.month) + '月' + str(ti.day) + '日 ' + str(ti.hour) + '时' + str(ti.minute) + '分'
time_file = str(ti.month) + '月' + str(ti.day) + '日 ' + str(ti.hour) + '时'
dest_filename='本科生院'+time_hot+'.xlsx'
wb=openpyxl.Workbook()
ws=wb.active
height=25.0
ws.cell(row=1, column=1, value='发布后先')
ws.cell(row=1, column=2, value='标题')
ws.cell(row=1,column=3,value='链接')
ws.cell(row=1,column=4,value='发布时间')
for i in range(len(biaoti)):
    ws.cell(row=i+2, column=1, value=i+1) 
    ws.cell(row=i+2, column=2, value=biaoti[i])
    ws.cell(row=i+2,column=3,value=f"https://www.bkjx.sdu.edu.cn/{lianjie[i]}")
    ws.cell(row=i+2,column=4,value=shijian[i])
ws.column_dimensions['B'].width = 130
ws.column_dimensions['D'].width = 15
for j in range(1,17):
    ws.row_dimensions[j].height = 30
wb.save(filename=dest_filename)
print('yes')